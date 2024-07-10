# how to process spreadsheets
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    # loading our workbook
    wb = xl.load_workbook(filename)
    # getting a reference to the first sheet
    sheet = wb['Sheet1']

    '''how to access a cell
    # get the coordinates of the cell
    # column then row
    cell = sheet['a1']
    # or use the .cell() method of the sheet object
    # pass in the row and column
    cell = sheet.cell(1,1)
    # shows transaction_id in console
    print(cell.value)
    
    # to determine how many rows in the spreadsheet
    # shows 4, so we need to generate a for loop that would generate the numbers 1-4
    print(sheet.max_row)'''

    for row in range(2, sheet.max_row + 1):
        price_cell = sheet.cell(row, 3)
        corrected_price = price_cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # use Reference class to select a range of values
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    # add chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # read docu for openpyxl to see different visualizations

    wb.save(filename)

